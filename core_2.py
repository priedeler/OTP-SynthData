import pandas as pd
import random
import numpy as np
from mimesis import Finance
from mimesis.locales import Locale
from datetime import datetime

def generate_company_name(genre):
    finance = Finance(locale=Locale.EN)
    name = finance.company()
    if genre == "Tech": name += " Technologies"
    elif genre == "Health": name += " Healthcare"
    elif genre == "Food": name += " Foods"
    elif genre == "Logistics": name += " Logistics"
    return name

def generate_materials(num_materials, brands=None, mat_classes=None, year=2025):
    # 1. TPMD_MaterialTypeClass (Static Master Data)
    if mat_classes is None:
        mat_classes_list = [
            {"Valuation Class": "OMP", "Type of Sales": "Intercompany & 3P", "Description": "Own Manufactured Products"},
            {"Valuation Class": "MER", "Type of Sales": "3P", "Description": "Merchandise"},
            {"Valuation Class": "RAW", "Type of Sales": "Intercompany", "Description": "Raw Materials"}
        ]
        df_mat_class = pd.DataFrame(mat_classes_list)
    else:
        df_mat_class = mat_classes.copy()
    
    # 2. TPMD_Material (Dynamic Master Data)
    materials = []
    if brands is None:
        brands = ["AlphaTech", "BetaMed", "GammaFoods", "DeltaLogistics", "NexGen", "CoreSystems"]
    
    for i in range(1, num_materials + 1):
        mat_id = f"MAT-{str(i).zfill(4)}"
        brand = random.choice(brands) if isinstance(brands, list) and brands else "Generic"
        
        raw_price = round(random.uniform(5.0, 50.0), 2)
        ic_price = round(raw_price * random.uniform(1.15, 1.40), 2)
        mer_price = round(raw_price * random.uniform(1.10, 1.30), 2)
        tp_price = round(max(ic_price, mer_price) * random.uniform(1.80, 3.50), 2)
        
        qty = random.randint(100, 5000)
        
        materials.append({
            "Material": mat_id,
            "Brand": brand,
            "Raw Material Price": raw_price,
            "IC Sales Price": ic_price,
            "MER Material Price": mer_price,
            "3P Sales Price": tp_price,
            "Column1": "",
            f"Quantities for Jan {year}": qty
        })
        
    return df_mat_class, pd.DataFrame(materials)

def generate_master_data():
    pnl_data = [
        {"P&L line": "Sales - IC", "MPR Account": "400000"},
        {"P&L line": "Sales - 3P", "MPR Account": "410000"},
        {"P&L line": "COGS - IC", "MPR Account": "500000"},
        {"P&L line": "COGS - 3P", "MPR Account": "510000"},
        {"P&L line": "Marketing", "MPR Account": "600000"},
        {"P&L line": "R&D", "MPR Account": "610000"},
        {"P&L line": "General Administration", "MPR Account": "620000"},
        {"P&L line": "Logistics", "MPR Account": "630000"}
    ]
    df_pnl = pd.DataFrame(pnl_data)

    segments = [
        {"TP Segment key": "DIST-T", "TP Segment description": "Distributor - TNMM"},
        {"TP Segment key": "DIST-R", "TP Segment description": "Distributor - RPM"},
        {"TP Segment key": "CM-T", "TP Segment description": "Routine Manufacturer"},
        {"TP Segment key": "CM-C", "TP Segment description": "Cost Plus Manufacturer"},
        {"TP Segment key": "PRIN", "TP Segment description": "IP Principal"},
        {"TP Segment key": "SERV", "TP Segment description": "Service Provider"},
        {"TP Segment key": "CUP", "TP Segment description": "Commodity Trader"}
    ]
    df_segments = pd.DataFrame(segments)
    
    return df_pnl, df_segments

def generate_config_data(companies_df, segments_df, role_counts=None, custom_benchmarks=None, year=2025):
    tp_segments = []
    
    if role_counts:
        roles_pool = (
            ["IP Principal"] * role_counts.get("principals", 2) +
            ["Distributor - TNMM", "Distributor - RPM"][:role_counts.get("distributors", 2)] +
            ["Routine Manufacturer", "Cost Plus Manufacturer"][:role_counts.get("manufacturers", 2)] +
            ["Service Provider"] * role_counts.get("service_providers", 1) +
            ["Commodity Trader"] * max(0, role_counts.get("traders", 0))
        )
        while len(roles_pool) < len(companies_df):
            roles_pool.append("Distributor - TNMM")
        roles_pool = roles_pool[:len(companies_df)]
    else:
        segments = segments_df["TP Segment description"].tolist()
        roles_pool = [random.choice(segments) for _ in range(len(companies_df))]

    for i, (_, row) in enumerate(companies_df.iterrows()):
        tp_segments.append({
            "Company Code": row["Company Code"],
            "#": "",
            "Pri": 1,
            "TP Segment": roles_pool[i],
            "Column1": "",
            "Type of Sales": "3P",
            "Valuation Class": "OMP",
            "Trading Partner?": "No",
            "Material Number": "*",
            "Valid from": f"01.01.{year}",
            "Valid to": f"31.12.{year}"
        })
    df_c_tp_segment = pd.DataFrame(tp_segments)

    if custom_benchmarks is not None:
        df_benchmark = custom_benchmarks.copy()
    else:
        benchmarks = [
            {"TP Function": "Distributor - TNMM", "TP Method": "TNMM", "Price Setting": "Target", "PLI Name": "OM", "PLI Formula": "EBIT/Sales", "Q1": 0.02, "Median": 0.035, "Q3": 0.05, "Valid from": f"01.01.{year}", "Valid to": f"31.12.{year}", "User ID": "SYS"},
            {"TP Function": "Distributor - RPM", "TP Method": "Resale Minus", "Price Setting": "Target", "PLI Name": "Gross Margin", "PLI Formula": "Gross Profit/Sales", "Q1": 0.15, "Median": 0.20, "Q3": 0.25, "Valid from": f"01.01.{year}", "Valid to": f"31.12.{year}", "User ID": "SYS"},
            {"TP Function": "Routine Manufacturer", "TP Method": "TNMM", "Price Setting": "Target", "PLI Name": "NCP", "PLI Formula": "EBIT/Total Cost", "Q1": 0.05, "Median": 0.075, "Q3": 0.10, "Valid from": f"01.01.{year}", "Valid to": f"31.12.{year}", "User ID": "SYS"},
            {"TP Function": "Service Provider", "TP Method": "TNMM", "Price Setting": "Target", "PLI Name": "Mark-up", "PLI Formula": "EBIT/Total Cost", "Q1": 0.03, "Median": 0.05, "Q3": 0.07, "Valid from": f"01.01.{year}", "Valid to": f"31.12.{year}", "User ID": "SYS"},
            {"TP Function": "Cost Plus Manufacturer", "TP Method": "Cost Plus", "Price Setting": "Target", "PLI Name": "Gross Mark-up", "PLI Formula": "Gross Profit/COGS", "Q1": 0.10, "Median": 0.15, "Q3": 0.20, "Valid from": f"01.01.{year}", "Valid to": f"31.12.{year}", "User ID": "SYS"},
            {"TP Function": "Commodity Trader", "TP Method": "CUP", "Price Setting": "Exact", "PLI Name": "Price", "PLI Formula": "Internal vs External", "Q1": 0.0, "Median": 0.0, "Q3": 0.0, "Valid from": f"01.01.{year}", "Valid to": f"31.12.{year}", "User ID": "SYS"},
            {"TP Function": "IP Principal", "TP Method": "PSM", "Price Setting": "Split", "PLI Name": "Residual", "PLI Formula": "OPEX Allocation Key", "Q1": 0.0, "Median": 0.0, "Q3": 0.0, "Valid from": f"01.01.{year}", "Valid to": f"31.12.{year}", "User ID": "SYS"}
        ]
        df_benchmark = pd.DataFrame(benchmarks)

    indirect_alloc = []
    pnl_lines = ["Marketing", "R&D", "General Administration", "Logistics"]
    for line in pnl_lines:
        indirect_alloc.append({
            "P&Ll line Item": line,
            "Account": "*",
            "Allocation Key": random.choice(["Headcount", "Revenue", "Floor Space"]),
            "User ID": "SYS"
        })
    df_indirect_alloc = pd.DataFrame(indirect_alloc)

    return df_c_tp_segment, df_benchmark, df_indirect_alloc

def generate_transactions(companies_df, materials_df, pnl_df, num_transactions, df_c_tp_segment, year=2025, qty_max=100, service_fee_min=5000, service_fee_max=25000, opex_min=0.08, opex_max=0.35):
    sales_tx = []
    company_codes = companies_df["Company Code"].tolist()
    
    co_to_country = dict(zip(companies_df["Company Code"], companies_df["Country Key"]))
    co_to_name = dict(zip(companies_df["Company Code"], companies_df["Company Name"]))
    co_to_currency = dict(zip(companies_df["Company Code"], companies_df["Co Currency"]))
    co_to_segment = dict(zip(df_c_tp_segment["Company Code"], df_c_tp_segment["TP Segment"]))
    
    principals = df_c_tp_segment[df_c_tp_segment["TP Segment"] == "IP Principal"]["Company Code"].tolist()
    distributors = df_c_tp_segment[df_c_tp_segment["TP Segment"].str.contains("Distributor")]["Company Code"].tolist()
    manufacturers = df_c_tp_segment[df_c_tp_segment["TP Segment"].str.contains("Manufacturer")]["Company Code"].tolist()
    service_providers = df_c_tp_segment[df_c_tp_segment["TP Segment"] == "Service Provider"]["Company Code"].tolist()
    commodity_traders = df_c_tp_segment[df_c_tp_segment["TP Segment"] == "Commodity Trader"]["Company Code"].tolist()
    
    ic_sales_acc = "400000"
    tp_sales_acc = "410000"
    ic_cogs_acc = "500000"
    tp_cogs_acc = "510000"
    
    revenue_tracker = {code: 0 for code in company_codes}
    
    # Speed Optimization: Convert materials DataFrame to a list of dictionaries
    materials_list = materials_df.to_dict('records')

    for flow_id in range(num_transactions):
        material = random.choice(materials_list)
        qty = random.randint(1, int(qty_max))
        month = random.randint(1, 12)
        period = f"{str(month).zfill(2)}.{year}"

        # --- RESILIENT ROUTING ENGINE ---
        current_buyer = None
        current_price = None
        product_sold_to_external = False
        
        # A. COMMODITY TRADER LEG (Optional raw material supply chain leg)
        trader_seller = None
        target_manufacturer = None
        if commodity_traders and manufacturers:
            trader_seller = random.choice(commodity_traders)
            target_manufacturer = random.choice(manufacturers)
            
            price_sales = material["IC Sales Price"]
            price_cogs = material["Raw Material Price"]
            revenue = round(qty * price_sales, 2)
            cogs = round(qty * price_cogs, 2)
            
            # Trader sells Raw Materials to Manufacturer (IC Sales)
            sales_tx.append({
                "Company Code": trader_seller,
                "Company": co_to_name[trader_seller],
                "Country Code": co_to_country[trader_seller],
                "Country": co_to_country[trader_seller],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account Sales": ic_sales_acc,
                "GL Description Sales": "Sales",
                "GL Account COGS": tp_cogs_acc,
                "GL Description COGS": "COGS",
                "MaterialNumber": material["Material"],
                "Brand": material["Brand"],
                "BusinessType": "Sales",
                "ValuationClass": "RAW",
                "TradingPartner": target_manufacturer,
                "Trading Partner": target_manufacturer,
                "Trading Partner Region": "Global",
                "TypeOfSales": "IC",
                "RUNIT": co_to_currency[trader_seller],
                "GlobalCurrency": "EUR",
                "Price Sales": price_sales,
                "Price COGS": price_cogs,
                "Total Sales": qty,
                "Total Amount Sales": revenue,
                "Total Amount COGS": -cogs,
                "Column": ""
            })
            revenue_tracker[trader_seller] += revenue
            
        # B. MANUFACTURING LEG
        if manufacturers:
            seller = target_manufacturer if (trader_seller is not None) else random.choice(manufacturers)
            
            if principals and distributors:
                if random.random() < 0.20:
                    buyer = random.choice(distributors)
                else:
                    buyer = random.choice(principals)
                type_sales = "IC"
            elif principals:
                buyer = random.choice(principals)
                type_sales = "IC"
            elif distributors:
                buyer = random.choice(distributors)
                type_sales = "IC"
            else:
                buyer = "EXTERNAL"
                type_sales = "3P"
            
            price_sales = material["IC Sales Price"] if buyer != "EXTERNAL" else material["3P Sales Price"]
            price_cogs = material["IC Sales Price"] if (trader_seller is not None) else material["Raw Material Price"]
            cogs_type = "IC" if (trader_seller is not None) else "3P"
            
            revenue = round(qty * price_sales, 2)
            cogs = round(qty * price_cogs, 2)
            
            sales_tx.append({
                "Company Code": seller,
                "Company": co_to_name[seller],
                "Country Code": co_to_country[seller],
                "Country": co_to_country[seller],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account Sales": ic_sales_acc if type_sales == "IC" else tp_sales_acc,
                "GL Description Sales": "Sales",
                "GL Account COGS": ic_cogs_acc if cogs_type == "IC" else tp_cogs_acc,
                "GL Description COGS": "COGS",
                "MaterialNumber": material["Material"],
                "Brand": material["Brand"],
                "BusinessType": "Sales",
                "ValuationClass": "OMP",
                "TradingPartner": buyer,
                "Trading Partner": buyer,
                "Trading Partner Region": "Global",
                "TypeOfSales": type_sales,
                "RUNIT": co_to_currency[seller],
                "GlobalCurrency": "EUR",
                "Price Sales": price_sales,
                "Price COGS": price_cogs,
                "Total Sales": qty,
                "Total Amount Sales": revenue,
                "Total Amount COGS": -cogs,
                "Column": ""
            })
            revenue_tracker[seller] += revenue
            if buyer != "EXTERNAL":
                current_buyer = buyer
                current_price = price_sales
            else:
                current_buyer = None
                current_price = None
                product_sold_to_external = True

        # C. PRINCIPAL LEG
        if principals and not product_sold_to_external:
            if current_buyer is not None and current_buyer not in principals:
                pass # Bypassed Principal
            else:
                seller = current_buyer if current_buyer else random.choice(principals)
                if distributors and random.random() < 0.80:
                    buyer = random.choice(distributors)
                    type_sales = "IC"
                    price_sales = material["MER Material Price"]
                else:
                    buyer = "EXTERNAL"
                    type_sales = "3P"
                    price_sales = material["3P Sales Price"]
                
                price_cogs = current_price if current_price is not None else material["Raw Material Price"]
                cogs_type = "IC" if current_price is not None else "3P"
                
                revenue = round(qty * price_sales, 2)
                cogs = round(qty * price_cogs, 2)
                
                sales_tx.append({
                    "Company Code": seller,
                    "Company": co_to_name[seller],
                    "Country Code": co_to_country[seller],
                    "Country": co_to_country[seller],
                    "Region CoCo": "Europe",
                    "Year": year,
                    "PeriodRange": period,
                    "GL Account Sales": ic_sales_acc if type_sales == "IC" else tp_sales_acc,
                    "GL Description Sales": "Sales",
                    "GL Account COGS": ic_cogs_acc if cogs_type == "IC" else tp_cogs_acc,
                    "GL Description COGS": "COGS",
                    "MaterialNumber": material["Material"],
                    "Brand": material["Brand"],
                    "BusinessType": "Sales",
                    "ValuationClass": "OMP",
                    "TradingPartner": buyer,
                    "Trading Partner": buyer,
                    "Trading Partner Region": "Global",
                    "TypeOfSales": type_sales,
                    "RUNIT": co_to_currency[seller],
                    "GlobalCurrency": "EUR",
                    "Price Sales": price_sales,
                    "Price COGS": price_cogs,
                    "Total Sales": qty,
                    "Total Amount Sales": revenue,
                    "Total Amount COGS": -cogs,
                    "Column": ""
                })
                revenue_tracker[seller] += revenue
                if buyer != "EXTERNAL":
                    current_buyer = buyer
                    current_price = price_sales
                else:
                    current_buyer = None
                    current_price = None
                    product_sold_to_external = True

        # D. DISTRIBUTION LEG
        if distributors and not product_sold_to_external:
            if current_buyer is not None and current_buyer not in distributors:
                pass # Should not happen unless bypassed distributor
            else:
                seller = current_buyer if current_buyer else random.choice(distributors)
                buyer = "EXTERNAL"
                
                price_cogs = current_price if current_price is not None else material["MER Material Price"]
                price_sales = material["3P Sales Price"]
                cogs_type = "IC" if current_price is not None else "3P"
                
                revenue = round(qty * price_sales, 2)
                cogs = round(qty * price_cogs, 2)
                
                sales_tx.append({
                "Company Code": seller,
                "Company": co_to_name[seller],
                "Country Code": co_to_country[seller],
                "Country": co_to_country[seller],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account Sales": tp_sales_acc,
                "GL Description Sales": "Sales",
                "GL Account COGS": ic_cogs_acc if cogs_type == "IC" else tp_cogs_acc,
                "GL Description COGS": "COGS",
                "MaterialNumber": material["Material"],
                "Brand": material["Brand"],
                "BusinessType": "Sales",
                "ValuationClass": "OMP",
                "TradingPartner": buyer,
                "Trading Partner": buyer,
                "Trading Partner Region": "Global",
                "TypeOfSales": "3P",
                "RUNIT": co_to_currency[seller],
                "GlobalCurrency": "EUR",
                "Price Sales": price_sales,
                "Price COGS": price_cogs,
                "Total Sales": qty,
                "Total Amount Sales": revenue,
                "Total Amount COGS": -cogs,
                "Column": ""
            })
            revenue_tracker[seller] += revenue

        # E. SERVICE PROVIDERS (Symmetric IC Accounting)
        if service_providers and random.random() < 0.20:
            seller = random.choice(service_providers)
            buyer = random.choice(principals) if principals else (random.choice(distributors) if distributors else random.choice(company_codes))
            
            service_fee = round(random.uniform(service_fee_min, service_fee_max), 2)
            
            # Seller Revenue Row
            sales_tx.append({
                "Company Code": seller,
                "Company": co_to_name[seller],
                "Country Code": co_to_country[seller],
                "Country": co_to_country[seller],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account Sales": ic_sales_acc,
                "GL Description Sales": "Service Fee",
                "GL Account COGS": ic_cogs_acc,
                "GL Description COGS": "Cost of Services",
                "MaterialNumber": "*",
                "Brand": "*",
                "BusinessType": "Service",
                "ValuationClass": "*",
                "TradingPartner": buyer,
                "Trading Partner": buyer,
                "Trading Partner Region": "Global",
                "TypeOfSales": "IC",
                "RUNIT": co_to_currency[seller],
                "GlobalCurrency": "EUR",
                "Price Sales": service_fee,
                "Price COGS": 0,
                "Total Sales": 1,
                "Total Amount Sales": service_fee,
                "Total Amount COGS": 0,
                "Column": ""
            })
            revenue_tracker[seller] += service_fee

            # Buyer Expense Row
            sales_tx.append({
                "Company Code": buyer,
                "Company": co_to_name[buyer],
                "Country Code": co_to_country[buyer],
                "Country": co_to_country[buyer],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account Sales": ic_sales_acc,
                "GL Description Sales": "Service Fee",
                "GL Account COGS": ic_cogs_acc,
                "GL Description COGS": "Purchased Services",
                "MaterialNumber": "*",
                "Brand": "*",
                "BusinessType": "Service",
                "ValuationClass": "*",
                "TradingPartner": seller,
                "Trading Partner": seller,
                "Trading Partner Region": "Global",
                "TypeOfSales": "IC",
                "RUNIT": co_to_currency[buyer],
                "GlobalCurrency": "EUR",
                "Price Sales": 0,
                "Price COGS": service_fee,
                "Total Sales": 0,
                "Total Amount Sales": 0,
                "Total Amount COGS": -service_fee,
                "Column": ""
            })

    # OPEX transactions (Scaled to Revenue)
    opex_tx = []
    opex_accounts = [("600000", "Marketing"), ("610000", "R&D"), ("620000", "General Admin"), ("630000", "Logistics")]
    
    for code, revenue in revenue_tracker.items():
        segment = co_to_segment.get(code, "Unknown")
        # Minimum OPEX if no revenue
        annual_opex = revenue * random.uniform(opex_min, opex_max) if revenue > 0 else random.uniform(50000, 200000)
            
        amount_per_entry = annual_opex / (len(opex_accounts) * 12)

        for opex_flow_id, (acc_code, acc_desc) in enumerate(opex_accounts):
            for month in range(1, 13):
                period = f"{str(month).zfill(2)}.{year}"
                monthly_amount = round(amount_per_entry * random.uniform(0.85, 1.15), 2)
                
                if segment == "Service Provider" and principals:
                    partner = random.choice(principals)
                else:
                    partner = "VARIOUS"
                    
                opex_tx.append({
                    "Company Code": code,
                    "Company": co_to_name[code],
                    "Country Code": co_to_country[code],
                    "Country": co_to_country[code],
                    "Region CoCo": "Europe",
                    "Year": year,
                    "PeriodRange": period,
                    "GL Account": acc_code,
                    "GL Description": acc_desc,
                    "MaterialNumber": "*",
                    "Brand": "*",
                    "BusinessType": "OPEX",
                    "ValuationClass": "*",
                    "TradingPartner": partner,
                    "Trading Partner": partner,
                    "Trading": partner,
                    "TypeOfSales": "OPEX",
                    "RUNIT": co_to_currency[code],
                    "GlobalCurrency": "EUR",
                    "Price": 0,
                    "Total Sales": 0,
                    "Total Amount": monthly_amount
                })

    return pd.DataFrame(sales_tx), pd.DataFrame(opex_tx)


def get_demo_scenario(year=2025):
    """
    Returns a hardcoded companies_df and df_c_tp_segment representing 
    the exact supply chain scenario provided in the diagram.
    """
    companies = [
        {"Company Code": "Co00", "Company Name": generate_company_name("General"), "City": "Zurich", "Country Name": "Switzerland", "Region": "Europe", "Country Key": "CHE", "Co Currency": "CHF", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "IP Principal", "lat": 47.3769, "lon": 8.5417},
        {"Company Code": "Co01-M", "Company Name": generate_company_name("General") + " Mfg", "City": "Paris", "Country Name": "France", "Region": "Europe", "Country Key": "FRA", "Co Currency": "EUR", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Routine Manufacturer", "lat": 48.8566, "lon": 2.3522},
        {"Company Code": "Co02", "Company Name": generate_company_name("General"), "City": "Warsaw", "Country Name": "Poland", "Region": "Europe", "Country Key": "POL", "Co Currency": "PLN", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Routine Manufacturer", "lat": 52.2297, "lon": 21.0122},
        {"Company Code": "Co03", "Company Name": generate_company_name("General"), "City": "Shanghai", "Country Name": "China", "Region": "Asia Pacific", "Country Key": "CHN", "Co Currency": "CNY", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Routine Manufacturer", "lat": 31.2304, "lon": 121.4737},
        {"Company Code": "Co01-D", "Company Name": generate_company_name("General") + " Dist", "City": "Lyon", "Country Name": "France", "Region": "Europe", "Country Key": "FRA", "Co Currency": "EUR", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 45.7640, "lon": 4.8357},
        {"Company Code": "Co04", "Company Name": generate_company_name("General"), "City": "Berlin", "Country Name": "Germany", "Region": "Europe", "Country Key": "DEU", "Co Currency": "EUR", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 52.5200, "lon": 13.4050},
        {"Company Code": "Co04-X", "Company Name": generate_company_name("General") + " XCZ", "City": "Munich", "Country Name": "Germany", "Region": "Europe", "Country Key": "DEU", "Co Currency": "EUR", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 48.1351, "lon": 11.5820},
        {"Company Code": "Co05", "Company Name": generate_company_name("General"), "City": "Oslo", "Country Name": "Norway", "Region": "Europe", "Country Key": "NOR", "Co Currency": "NOK", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 59.9139, "lon": 10.7522},
        {"Company Code": "Co11", "Company Name": generate_company_name("General"), "City": "New York", "Country Name": "United States", "Region": "Americas", "Country Key": "USA", "Co Currency": "USD", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 40.7128, "lon": -74.0060},
        {"Company Code": "Co12", "Company Name": generate_company_name("General"), "City": "Toronto", "Country Name": "Canada", "Region": "Americas", "Country Key": "CAN", "Co Currency": "CAD", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 43.6510, "lon": -79.3470},
        {"Company Code": "Co23", "Company Name": generate_company_name("General"), "City": "Dubai", "Country Name": "United Arab Emirates", "Region": "Middle East", "Country Key": "ARE", "Co Currency": "AED", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 25.2048, "lon": 55.2708},
        {"Company Code": "Co27", "Company Name": generate_company_name("General"), "City": "Tokyo", "Country Name": "Japan", "Region": "Asia Pacific", "Country Key": "JPN", "Co Currency": "JPY", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 35.6762, "lon": 139.6503},
        {"Company Code": "Co28", "Company Name": generate_company_name("General"), "City": "Mumbai", "Country Name": "India", "Region": "Asia Pacific", "Country Key": "IND", "Co Currency": "INR", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 19.0760, "lon": 72.8777},
        {"Company Code": "Co29", "Company Name": generate_company_name("General"), "City": "Singapore", "Country Name": "Singapore", "Region": "Asia Pacific", "Country Key": "SGP", "Co Currency": "SGD", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 1.3521, "lon": 103.8198},
        {"Company Code": "Co21", "Company Name": generate_company_name("General"), "City": "Bangkok", "Country Name": "Thailand", "Region": "Asia Pacific", "Country Key": "THA", "Co Currency": "THB", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 13.7563, "lon": 100.5018},
        {"Company Code": "Co24", "Company Name": generate_company_name("General"), "City": "Kuala Lumpur", "Country Name": "Malaysia", "Region": "Asia Pacific", "Country Key": "MYS", "Co Currency": "MYR", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 3.1390, "lon": 101.6869},
        {"Company Code": "Co33", "Company Name": generate_company_name("General"), "City": "Buenos Aires", "Country Name": "Argentina", "Region": "Americas", "Country Key": "ARG", "Co Currency": "ARS", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": -34.6037, "lon": -58.3816},
        {"Company Code": "Co34", "Company Name": generate_company_name("General"), "City": "São Paulo", "Country Name": "Brazil", "Region": "Americas", "Country Key": "BRA", "Co Currency": "BRL", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": -23.5505, "lon": -46.6333},
        {"Company Code": "Co35", "Company Name": generate_company_name("General"), "City": "Bogotá", "Country Name": "Colombia", "Region": "Americas", "Country Key": "COL", "Co Currency": "COP", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": 4.7110, "lon": -74.0721},
        {"Company Code": "Co46", "Company Name": generate_company_name("General"), "City": "Johannesburg", "Country Name": "South Africa", "Region": "Africa", "Country Key": "ZAF", "Co Currency": "ZAR", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": -26.2041, "lon": 28.0473},
        {"Company Code": "Co41", "Company Name": generate_company_name("General"), "City": "Luanda", "Country Name": "Angola", "Region": "Africa", "Country Key": "AGO", "Co Currency": "AOA", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": -8.8390, "lon": 13.2894},
        {"Company Code": "Co52", "Company Name": generate_company_name("General"), "City": "Sydney", "Country Name": "Australia", "Region": "Asia Pacific", "Country Key": "AUS", "Co Currency": "AUD", "Group Currency": "EUR", "Language Key": "EN", "Chart of Accounts": "COA", "TP Segment": "Distributor - TNMM", "lat": -33.8688, "lon": 151.2093}
    ]
    
    companies_df = pd.DataFrame(companies)
    
    tp_segments = []
    for i, row in companies_df.iterrows():
        tp_segments.append({
            "Company Code": row["Company Code"],
            "#": "",
            "Pri": 1,
            "TP Segment": row["TP Segment"],
            "Column1": "",
            "Type of Sales": "3P",
            "Valuation Class": "OMP",
            "Trading Partner?": "No",
            "Material Number": "*",
            "Valid from": f"01.01.{year}",
            "Valid to": f"31.12.{year}"
        })
    df_c_tp_segment = pd.DataFrame(tp_segments)
    
    # Remove the TP Segment column from companies_df to match original schema
    companies_df = companies_df.drop(columns=["TP Segment"])
    
    return companies_df, df_c_tp_segment

def generate_tp_adjustments(sales_tx, opex_tx, companies_df, p_total, df_c_tp_segment, year=2025):
    """
    Generate symmetric operational TP adjustment line items formatted exactly like standard opex/sales rows.
    """
    adj_entries = []
    
    co_to_country = dict(zip(companies_df["Company Code"], companies_df["Country Key"]))
    co_to_name = dict(zip(companies_df["Company Code"], companies_df["Company Name"]))
    co_to_currency = dict(zip(companies_df["Company Code"], companies_df["Co Currency"]))
    
    principals = df_c_tp_segment[df_c_tp_segment["TP Segment"] == "IP Principal"]["Company Code"].tolist()
    
    # Pre-calculate principal weights based on OPEX
    is_principal = p_total["TP Segment"] == "IP Principal"
    principal_weights = {}
    if is_principal.any():
        total_principal_opex = p_total.loc[is_principal, "OPEX"].sum()
        for _, p_row in p_total[is_principal].iterrows():
            p_code = p_row["Company Code"]
            if total_principal_opex > 0:
                principal_weights[p_code] = p_row["OPEX"] / total_principal_opex
            else:
                principal_weights[p_code] = 1.0 / is_principal.sum()
    
    for _, row in p_total.iterrows():
        code = row["Company Code"]
        tp_adj = row["TP Adjustment"]
        
        if abs(tp_adj) < 1.0 or row["TP Segment"] == "IP Principal":
            continue
            
        if not principals:
            partner = "EXTERNAL"
            period = f"12.{year}"
            
            # Routine Entity posting
            adj_entries.append({
                "Company Code": code,
                "Company": co_to_name[code],
                "Country Code": co_to_country[code],
                "Country": co_to_country[code],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account": "490000",
                "GL Description": "Transfer Pricing Adjustment",
                "MaterialNumber": "*",
                "Brand": "*",
                "BusinessType": "TP_ADJ",
                "ValuationClass": "*",
                "TradingPartner": partner,
                "Trading Partner": partner,
                "Trading": partner,
                "TypeOfSales": "OPEX",
                "RUNIT": co_to_currency[code],
                "GlobalCurrency": "EUR",
                "Price": 0,
                "Total Sales": 0,
                "Total Amount": round(tp_adj, 2)
            })
            continue

        # Split the tp_adj among all principals according to their weights
        for partner, weight in principal_weights.items():
            split_adj = tp_adj * weight
            if abs(split_adj) < 0.01: continue
            
            period = f"12.{year}"
            
            # Routine Entity posting
            adj_entries.append({
                "Company Code": code,
                "Company": co_to_name[code],
                "Country Code": co_to_country[code],
                "Country": co_to_country[code],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account": "490000",
                "GL Description": "Transfer Pricing Adjustment",
                "MaterialNumber": "*",
                "Brand": "*",
                "BusinessType": "TP_ADJ",
                "ValuationClass": "*",
                "TradingPartner": partner,
                "Trading Partner": partner,
                "Trading": partner,
                "TypeOfSales": "OPEX",
                "RUNIT": co_to_currency[code],
                "GlobalCurrency": "EUR",
                "Price": 0,
                "Total Sales": 0,
                "Total Amount": round(split_adj, 2)
            })
            
            # Symmetric Principal posting
            adj_entries.append({
                "Company Code": partner,
                "Company": co_to_name[partner],
                "Country Code": co_to_country[partner],
                "Country": co_to_country[partner],
                "Region CoCo": "Europe",
                "Year": year,
                "PeriodRange": period,
                "GL Account": "590000",
                "GL Description": "Transfer Pricing Cost Offset",
                "MaterialNumber": "*",
                "Brand": "*",
                "BusinessType": "TP_ADJ",
                "ValuationClass": "*",
                "TradingPartner": code,
                "Trading Partner": code,
                "Trading": code,
                "TypeOfSales": "OPEX",
                "RUNIT": co_to_currency[partner],
                "GlobalCurrency": "EUR",
                "Price": 0,
                "Total Sales": 0,
                "Total Amount": round(-split_adj, 2)
            })
            
    return pd.DataFrame(adj_entries)

def calculate_allocations(sales_tx, opex_tx, companies_df, df_benchmark, df_c_tp_segment, return_all=False):
    # Aggregated view
    p_total = sales_tx.groupby(["Company Code", "Company", "Country Code", "Country", "Co Currency" if "Co Currency" in sales_tx.columns else "RUNIT"]).agg({
        "Total Amount Sales": "sum",
        "Total Amount COGS": "sum",
        "Total Sales": "sum"
    }).reset_index()
    
    p_total.rename(columns={"Total Sales": "Sales Quantity", "Total Amount Sales": "Revenue", "Total Amount COGS": "COGS"}, inplace=True)
    p_total["COGS"] = p_total["COGS"].abs()
    
    opex_sum = opex_tx.groupby("Company Code")["Total Amount"].sum().reset_index()
    opex_sum.rename(columns={"Total Amount": "OPEX"}, inplace=True)
    p_total = p_total.merge(opex_sum, on="Company Code", how="left").fillna(0)
    p_total["OPEX"] = p_total["OPEX"].abs()
    
    p_total = p_total.merge(df_c_tp_segment[["Company Code", "TP Segment"]], on="Company Code", how="left")
    p_total = p_total.merge(df_benchmark[["TP Function", "Median", "PLI Name", "TP Method"]], left_on="TP Segment", right_on="TP Function", how="left")
    p_total.rename(columns={"Median": "Target Margin"}, inplace=True)
    
    # Calculate Base Profits
    p_total["Gross Profit"] = p_total["Revenue"] - p_total["COGS"]
    p_total["Preliminary Operating Profit"] = p_total["Gross Profit"] - p_total["OPEX"]
    
    is_principal = p_total["TP Method"] == "PSM"
    is_routine = ~is_principal

    p_total["Target OP"] = 0.0
    p_total["Target Gross Profit"] = 0.0
    
    # 1. Resale Price Method (RPM) -> Target Gross = Revenue * Target Margin
    rpm_mask = p_total["TP Method"] == "Resale Minus"
    p_total.loc[rpm_mask, "Target Gross Profit"] = p_total["Revenue"] * p_total["Target Margin"]
    p_total.loc[rpm_mask, "Target OP"] = p_total["Target Gross Profit"] - p_total["OPEX"]
    
    # 2. Traditional Cost Plus (CPL) -> Target Gross = COGS * Target Margin
    cpl_mask = p_total["TP Method"] == "Cost Plus"
    p_total.loc[cpl_mask, "Target Gross Profit"] = p_total["COGS"] * p_total["Target Margin"]
    p_total.loc[cpl_mask, "Target OP"] = p_total["Target Gross Profit"] - p_total["OPEX"]

    # 3. TNMM (OM and NCP) -> Target OP calculated directly
    tnmm_om_mask = (p_total["TP Method"] == "TNMM") & (p_total["PLI Name"] == "OM")
    p_total.loc[tnmm_om_mask, "Target OP"] = p_total["Revenue"] * p_total["Target Margin"]
    
    # NCP
    tnmm_ncp_mask = (p_total["TP Method"] == "TNMM") & (p_total["PLI Name"] == "NCP")
    p_total.loc[tnmm_ncp_mask, "Target OP"] = (p_total["COGS"] + p_total["OPEX"]) * p_total["Target Margin"]

    # 4. CUP -> Managed at the transactional level, no year-end OP adjustment needed
    cup_mask = p_total["TP Method"] == "CUP"
    p_total.loc[cup_mask, "Target OP"] = p_total["Preliminary Operating Profit"]

    # Calculate Adjustments for Routine Entities
    p_total["TP Adjustment"] = 0.0
    p_total.loc[is_routine, "TP Adjustment"] = p_total["Target OP"] - p_total["Preliminary Operating Profit"]
    
    # 5. Contribution Profit Split Method (PSM)
    total_routine_adjustment = p_total.loc[is_routine, "TP Adjustment"].sum()
    if is_principal.any():
        total_principal_opex = p_total.loc[is_principal, "OPEX"].sum()
        if total_principal_opex > 0:
            # Split based on OPEX Allocation Key (proxy for R&D/Substance)
            p_total.loc[is_principal, "TP Adjustment"] = (-total_routine_adjustment) * (p_total.loc[is_principal, "OPEX"] / total_principal_opex)
        else:
            # Fallback to even split
            p_total.loc[is_principal, "TP Adjustment"] = -total_routine_adjustment / is_principal.sum()
    
    # Finalize
    p_total["Final Operating Profit"] = p_total["Preliminary Operating Profit"] + p_total["TP Adjustment"]
    p_total.loc[is_principal, "Target OP"] = p_total["Final Operating Profit"]
    
    cols_to_round = ["Revenue", "COGS", "Gross Profit", "OPEX", "Target OP", "Preliminary Operating Profit", "TP Adjustment", "Final Operating Profit"]
    p_total[cols_to_round] = p_total[cols_to_round].round(2)
    
    extra_cols = ["Sales", "Sales IC", "Rebates", "Discounts", "Deductions", "Other Income"]
    for col in extra_cols:
        if col not in p_total.columns:
            p_total[col] = 0

    if return_all:
        # 1. P_Segmentation_Sales_COGS
        p_seg_sales = sales_tx.copy()
        p_seg_sales.rename(columns={"Region CoCo": "Region", "Trading Partner": "Trading Partner Country"}, inplace=True)
        p_seg_sales["TP Function"] = "Distributor"
        if "Column" in p_seg_sales.columns:
            p_seg_sales.drop(columns=["Column"], inplace=True)

        # 2. P_Segmentation_OPEX
        p_seg_opex = opex_tx.copy()
        p_seg_opex.rename(columns={
            "Region CoCo": "Region",
            "GL Account": "GL Account Sales",
            "GL Description": "GL Description Sales",
            "Total Amount": "Total Amount Sales",
            "Trading Partner": "Trading Partner Country",
            "Trading": "Trading Partner Region"
        }, inplace=True)
        p_seg_opex["CoCo"] = p_seg_opex["Company Code"]
        p_seg_opex["GL Account COGS"] = ""
        p_seg_opex["GL Description COGS"] = ""
        p_seg_opex["Total Amount COGS"] = 0
        p_seg_opex["Price Sales"] = 0
        p_seg_opex["Price COGS"] = 0
        p_seg_opex["TP Function"] = "Distributor"
        
        # 3. P_Direct Allocation
        p_direct = p_seg_opex.copy()
        
        # 4. P_Indirect Allocation
        p_indirect = opex_tx.copy()
        p_indirect.rename(columns={"Region CoCo": "Region", "Trading": "Trading Partner Region", "Trading Partner": "Trading Partner Country"}, inplace=True)
        p_indirect["CoCo"] = p_indirect["Company Code"]
        
        return p_seg_sales, p_seg_opex, p_direct, p_indirect, p_total

    return p_total

def save_report_to_template(template_path, output_path_or_stream, data_dict):
    """
    Load an Excel template from template_path, write the dataframes from data_dict
    to matching sheets while preserving column order and other sheets.
    This does NOT modify the original template file on disk.
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(template_path)
    
    for sheet_name, df in data_dict.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Read headers from row 1
            headers = [cell.value for cell in ws[1]]
            
            # Reindex dataframe to match template headers exactly
            df_reindexed = df.reindex(columns=headers)
            
            # Clear all rows below header
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
                
            # Append rows
            for row in df_reindexed.itertuples(index=False, name=None):
                # Convert NumPy NaNs to None so they render as empty cells in Excel
                row_list = [None if pd.isna(val) else val for val in row]
                ws.append(row_list)
                
    wb.save(output_path_or_stream)
